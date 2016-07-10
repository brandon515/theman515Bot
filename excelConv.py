import openpyxl as xl
import sqlite3

def main():
    values = []
    wb = xl.load_workbook("pashto.xlsx")
    ws = wb.active
    index = 0
    for x in xrange(1,ws.max_row):
        englishWord = ws.cell(row = x, column = 1)
        pashtoWord = ws.cell(row = x, column = 3)
        if englishWord.value != None and pashtoWord.value != None:
            values.append((index, englishWord.value, pashtoWord.value))
            print(values[index])
            index = index+1
    cmd = "INSERT INTO DICT (ID, ENGLISH, PASHTO) VALUES (?,?,?)"
    db = sqlite3.connect('pashto.db')
    cur = db.cursor()
    cur.executemany(cmd, values)
    db.commit()

if __name__ == "__main__":
    main()

